import unicodedata
from datetime import datetime, timedelta

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils.encoding import escape_uri_path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, NamedStyle
from openpyxl.utils import get_column_letter

from ..models import Student, Schedule
from ..serializers import StudentDetailSerializer, StudentRelativesSerializer

def export_student_schedule_to_excel(request, student_id):
    try:
        student = get_object_or_404(Student, id=student_id)
        student_serializer = StudentDetailSerializer(student)
        relatives_serializer = StudentRelativesSerializer(student)
        student_data = student_serializer.data
        relatives_data = relatives_serializer.data['relatives']
        schedules = Schedule.objects.filter(students=student).order_by('time_start')
        
        wb = Workbook()
        ws_info = wb.active
        ws_info.title = "Информация о обучающиемся"
        
        def write_cell(ws, row, col, value):
            cell = ws.cell(row=row, column=col)
            if isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = '0'
            elif isinstance(value, datetime):
                cell.value = value
                cell.number_format = 'dd.mm.yyyy'
            else:
                cell.value = str(value)
        
        row = 1
        for field, value in student_data.items():
            if field != 'Родственники':  # Пропускаем поле с родственниками
                write_cell(ws_info, row, 1, field)
                write_cell(ws_info, row, 2, value)
                row += 1
        
        # Создаем отдельный лист для родственников
        ws_relatives = wb.create_sheet(title="Родственники")
        
        # Заголовки для родственников
        relative_headers = ['ФИО', 'Пол', 'Отношение']
        for col, header in enumerate(relative_headers, start=1):
            write_cell(ws_relatives, 1, col, header)
        
        # Заполняем данные родственников
        for row, relative in enumerate(relatives_data, start=2):
            write_cell(ws_relatives, row, 1, relative['full_name'])
            write_cell(ws_relatives, row, 2, relative['gender_display'])
            write_cell(ws_relatives, row, 3, relative['relation_display'])
        
        # Создаем лист с расписанием
        ws_schedule = wb.create_sheet(title="Расписание")
        
        # Заголовки
        headers = ["Неделя", "Дата", "Время", "Кабинет", "Название", "Специалисты"]
        for col, header in enumerate(headers, start=1):
            write_cell(ws_schedule, 1, col, header)
        
        # Стили для заголовков
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        for ws in [ws_info, ws_relatives, ws_schedule]:
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = header_alignment
        
        # Заполняем расписание
        row = 2
        current_week_start = None
        for schedule in schedules:
            start_date = schedule.time_start.date()
            end_date = schedule.time_end.date()
            if current_week_start is None or start_date >= current_week_start + timedelta(days=7):
                current_week_start = start_date - timedelta(days=start_date.weekday())
                week_end = current_week_start + timedelta(days=6)
                week_str = f"{current_week_start.strftime('%d.%m.%Y')} - {week_end.strftime('%d.%m.%Y')}"
            
            # Используем время напрямую, без корректировки, так как Django должен учитывать часовой пояс
            adjusted_start_time = schedule.time_start
            adjusted_end_time = schedule.time_end
            
            write_cell(ws_schedule, row, 1, week_str)
            write_cell(ws_schedule, row, 2, f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}")
            write_cell(ws_schedule, row, 3, f"{adjusted_start_time.strftime('%H:%M')} - {adjusted_end_time.strftime('%H:%M')}")
            write_cell(ws_schedule, row, 4, schedule.room if schedule.room else "Не указана")
            write_cell(ws_schedule, row, 5, schedule.name)
            write_cell(ws_schedule, row, 6, ", ".join([specialist.full_name for specialist in schedule.specialists.all()]))
            row += 1

        
        # Автоматическая ширина столбцов
        for ws in [ws_info, ws_relatives, ws_schedule]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f'Расписаниe_{unicodedata.normalize("NFKD", student.full_name)}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{escape_uri_path(filename)}"'
        
        wb.save(response)
        
        return response
    
    except Student.DoesNotExist:
        return HttpResponse("Обучающийся не найден", status=404)
    except Exception as e:
        return HttpResponse(f"Произошла ошибка при создании файла: {str(e)}", status=400)

def export_all_students_to_excel(request):
    try:
        age_group = request.query_params.get('age_group')
        students = Student.objects.prefetch_related(
            Prefetch('relatives', queryset=Relative.objects.all())
        )
        
        if age_group in dict(Student.AGE_GROUP_CHOICES):
            students = students.filter(age_group=age_group)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Список обучающихся"
        
        # Создаем стиль с переносом текста
        wrap_style = NamedStyle(name="wrap_style")
        wrap_style.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Заголовки
        headers = [
            "Ф.И.О.", "Дата рождения", "Долги", "Примечание", "Адрес проживания / Адрес прописки",
            "Район", "Город / село", "Контактный телефон", "Установление недееспособности",
            "Отметка посещаемости", "Образование (школьное, ср.специальное)", "Доп.образование / Кружки",
            "Спорт", "Услуги соцзащиты в др. учреждениях", "Отказ от диагностики СП", "Отказ от фото",
            "№ Договора", "Дата договора", "№ ИППСУ", "Дата ИППСУ", "Справка 202Н",
            "Родитель/законный представитель/инвалидность", "№ Договора родителя", "Дата договора родителя",
            "№ ИППСУ родителя", "Дата ИППСУ родителя", "Справка 202Н дата выдачи родителя",
            "Группа инвалидности", "ПМПК (дата/номер)", "Заболевание", "№ИПРА", "Дата ИПРА",
            "Степень ограничений", "МСЭ № и дата", "Состав семьи", "Категория семьи", "Лагерь/лето",
            "Оплата услуг", "Лояльность (да/нет)", "Отказ от замены специалиста"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        
        # Заполнение данными
        for row, student in enumerate(students, start=2):
            relative = student.relatives.first()  # Получаем первого родственника
            
            ws.cell(row=row, column=1, value=student.full_name).style = wrap_style
            ws.cell(row=row, column=2, value=student.birth_date).style = wrap_style
            ws.cell(row=row, column=3, value=student.debts).style = wrap_style
            ws.cell(row=row, column=4, value=student.note).style = wrap_style
            ws.cell(row=row, column=5, value=f"{student.residence_address} / {student.registration_address}").style = wrap_style
            ws.cell(row=row, column=6, value=student.district).style = wrap_style
            ws.cell(row=row, column=7, value="Город" if student.is_city else "Село").style = wrap_style
            ws.cell(row=row, column=8, value=student.contact_phone).style = wrap_style
            ws.cell(row=row, column=9, value=dict(Student.CAPACITY_CHOICES).get(student.capacity_status, '')).style = wrap_style
            ws.cell(row=row, column=10, value=dict(Student.ATTENDANCE_CHOICES).get(student.attendance_status, '')).style = wrap_style
            ws.cell(row=row, column=11, value=student.has_education).style = wrap_style
            ws.cell(row=row, column=12, value=student.additional_education_name if student.has_additional_education else "Нет").style = wrap_style
            ws.cell(row=row, column=13, value=student.sports_name if student.has_sports else "Нет").style = wrap_style
            ws.cell(row=row, column=14, value="Да" if student.has_other_social_services else "Нет").style = wrap_style
            ws.cell(row=row, column=15, value="Отказ" if student.refused_diagnostics else "").style = wrap_style
            ws.cell(row=row, column=16, value="Отказ" if student.refused_photo else "").style = wrap_style
            ws.cell(row=row, column=17, value=student.contract_number).style = wrap_style
            ws.cell(row=row, column=18, value=student.contract_date).style = wrap_style
            ws.cell(row=row, column=19, value=student.ippsu_number).style = wrap_style
            ws.cell(row=row, column=20, value=student.ippsu_date).style = wrap_style
            ws.cell(row=row, column=21, value=student.certificate_202n).style = wrap_style
            ws.cell(row=row, column=22, value=relative.relation if relative else "").style = wrap_style
            ws.cell(row=row, column=23, value=student.parent_contract_number).style = wrap_style
            ws.cell(row=row, column=24, value=student.parent_contract_date).style = wrap_style
            ws.cell(row=row, column=25, value=student.parent_ippsu_number).style = wrap_style
            ws.cell(row=row, column=26, value=student.parent_ippsu_date).style = wrap_style
            ws.cell(row=row, column=27, value=student.parent_certificate_202n_date).style = wrap_style
            ws.cell(row=row, column=28, value=student.disability_group).style = wrap_style
            ws.cell(row=row, column=29, value=f"{student.pmpk_date}/{student.pmpk_number}").style = wrap_style
            ws.cell(row=row, column=30, value=student.disease).style = wrap_style
            ws.cell(row=row, column=31, value=student.ipra_number).style = wrap_style
            ws.cell(row=row, column=32, value=student.ipra_date).style = wrap_style
            ws.cell(row=row, column=33, value=student.limitation_degree).style = wrap_style
            ws.cell(row=row, column=34, value=f"{student.mse_number} {student.mse_date}").style = wrap_style
            ws.cell(row=row, column=35, value=dict(Student.FAMILY_COMPOSITION_CHOICES).get(student.family_composition, '')).style = wrap_style
            ws.cell(row=row, column=36, value=dict(Student.FAMILY_CATEGORY_CHOICES).get(student.family_category, '')).style = wrap_style
            ws.cell(row=row, column=37, value=student.summer_camp_note if student.has_summer_camp else "Нет").style = wrap_style
            ws.cell(row=row, column=38, value=dict(Student.PAYMENT_CHOICES).get(student.payment_type, '')).style = wrap_style
            ws.cell(row=row, column=39, value="Да" if student.loyalty else "Нет").style = wrap_style
            ws.cell(row=row, column=40, value="Да" if student.specialist_change_refusal else "Нет").style = wrap_style
        
        # Автоматическая ширина столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50) * 1.2  # Ограничиваем максимальную ширину
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=all_students.xlsx'
        
        wb.save(response)
        
        return response
    
    except Exception as e:
        return HttpResponse(f"Произошла ошибка при создании файла: {str(e)}", status=400)
    
    

from django.utils import timezone
from datetime import timedelta
from ..models import Specialist, Schedule


def export_specialist_schedule_to_excel(specialist, start_date, end_date):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Расписание специалиста"

        # Заголовки
        headers = ["Дата", "Время", "Кабинет", "Название", "Обучающиеся", "Присутствовали"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)

        # Получаем расписание специалиста
        schedules = Schedule.objects.filter(specialists=specialist).order_by('time_start')
        if start_date:
            schedules = schedules.filter(time_start__gte=start_date)
        if end_date:
            schedules = schedules.filter(time_end__lte=end_date)

        # Заполняем расписание
        row = 2
        current_date = None
        for schedule in schedules:
            try:
                adjusted_start_time = schedule.time_start
                adjusted_end_time = schedule.time_end

                # Если дата изменилась, добавляем пустую строку
                if current_date and current_date != adjusted_start_time.date():
                    row += 1

                current_date = adjusted_start_time.date()

                ws.cell(row=row, column=1, value=adjusted_start_time.strftime('%d.%m.%Y'))
                ws.cell(row=row, column=2, value=f"{adjusted_start_time.strftime('%H:%M')} - {adjusted_end_time.strftime('%H:%M')}")
                ws.cell(row=row, column=3, value=schedule.room.name if schedule.room else "Не указана")
                ws.cell(row=row, column=4, value=schedule.name)
                
                all_students = [student.full_name for student in schedule.students.all()]
                ws.cell(row=row, column=5, value=", ".join(all_students))
                
                # Добавляем информацию о присутствии
                present_students = [attendance.student.full_name for attendance in schedule.attendances.filter(is_present=True)]
                ws.cell(row=row, column=6, value=", ".join(present_students))
                
                row += 1
            except Exception as e:
                print(f"Ошибка при обработке строки {row}: {str(e)}")

        # Автоматическая ширина столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f'Расписание_{unicodedata.normalize("NFKD", specialist.full_name)}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{escape_uri_path(filename)}"'

        wb.save(response)

        return response

    except Exception as e:
        print(f"Ошибка при создании Excel-файла: {str(e)}")
        return HttpResponse(f"Произошла ошибка при создании файла: {str(e)}", status=500)

